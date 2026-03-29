import os
import csv
import json
import uuid
import shutil
import base64
import hashlib
import datetime
from flask import (Flask, render_template, request, redirect, url_for,
                   session, flash, jsonify, send_file, send_from_directory)
from werkzeug.utils import secure_filename

# Load .env file
try:
    from dotenv import load_dotenv
    load_dotenv(os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env"))
except ImportError:
    pass

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'liberty-emporium-secret-2026')

# ── Paths ────────────────────────────────────────────────────────────────────
BASE_DIR       = os.path.dirname(os.path.abspath(__file__))
DATA_DIR       = os.environ.get('DATA_DIR', BASE_DIR)
INVENTORY_FILE = os.path.join(DATA_DIR, 'inventory.csv')
UPLOAD_FOLDER  = os.path.join(DATA_DIR, 'uploads')
MUSIC_FOLDER   = os.path.join(DATA_DIR, 'music')
BACKUP_FOLDER  = os.path.join(DATA_DIR, 'backups')
ADS_FOLDER     = os.path.join(DATA_DIR, 'ads')
USERS_FILE     = os.path.join(DATA_DIR, 'users.json')
PENDING_FILE   = os.path.join(DATA_DIR, 'pending_users.json')

for d in [UPLOAD_FOLDER, BACKUP_FOLDER, ADS_FOLDER, MUSIC_FOLDER]:
    os.makedirs(d, exist_ok=True)

# ── Config ───────────────────────────────────────────────────────────────────
STORE_NAME    = 'Liberty Emporium & Thrift'
DEMO_MODE     = os.environ.get('DEMO_MODE', 'false').lower() == 'true'
CONTACT_EMAIL = os.environ.get('CONTACT_EMAIL', 'alexanderjay70@gmail.com')
ALLOWED_EXT   = {'jpg', 'jpeg', 'png', 'gif', 'webp'}
MAX_BACKUPS   = 20

CATEGORIES = ['Furniture','Electronics','Clothing','Jewelry','Home Decor',
              'Books','Kitchen','Toys','Tools','Collectibles','Art','Miscellaneous']
CONDITIONS = ['New','Like New','Good','Fair','Poor']
STATUSES   = ['Available','Sold','Reserved','Pending']

ADMIN_USER  = 'admin'
ADMIN_PASS  = os.environ.get('ADMIN_PASSWORD', 'admin123')
ADMIN_EMAIL = 'alexanderjay70@gmail.com'

# ── Gmail config for password reset emails ────────────────────────────────────
# Set these as environment variables on PythonAnywhere, or fill in directly.
# To get a Gmail App Password: myaccount.google.com → Security → 2-Step → App Passwords
GMAIL_ADDRESS  = os.environ.get('GMAIL_ADDRESS',  'emporiumandthrift@gmail.com')
GMAIL_APP_PASS = os.environ.get('GMAIL_APP_PASS', '')   # ← paste your App Password here
RESET_TOKEN_FILE = os.path.join(BASE_DIR, 'reset_tokens.json')

# ── Helpers ──────────────────────────────────────────────────────────────────
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXT

def hash_password(pw):
    return hashlib.sha256(pw.encode()).hexdigest()

def load_users():
    if not os.path.exists(USERS_FILE):
        return {}
    with open(USERS_FILE) as f:
        return json.load(f)

def save_users(users):
    with open(USERS_FILE, 'w') as f:
        json.dump(users, f, indent=2)

# ── Reset token helpers ───────────────────────────────────────────────────────
def load_reset_tokens():
    if not os.path.exists(RESET_TOKEN_FILE):
        return {}
    with open(RESET_TOKEN_FILE) as f:
        return json.load(f)

def save_reset_tokens(tokens):
    with open(RESET_TOKEN_FILE, 'w') as f:
        json.dump(tokens, f, indent=2)

def send_reset_email(to_addr, reset_url):
    import smtplib
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart
    msg = MIMEMultipart('alternative')
    msg['Subject'] = f'Password Reset – {STORE_NAME}'
    msg['From']    = GMAIL_ADDRESS
    msg['To']      = to_addr
    text = f"""Hi,

You requested a password reset for your {STORE_NAME} account.

Click the link below to set a new password (expires in 1 hour):

{reset_url}

If you did not request this, ignore this email – your password will not change.

– {STORE_NAME}
"""
    html = f"""<div style="font-family:Arial,sans-serif;max-width:520px;margin:0 auto;">
  <div style="background:#8B0000;padding:1.5rem 2rem;">
    <h2 style="color:#D4A017;margin:0;font-size:1.4rem;">🔑 Password Reset</h2>
    <p style="color:rgba(255,255,255,0.7);margin:0.3rem 0 0;font-size:0.9rem;">{STORE_NAME}</p>
  </div>
  <div style="background:#FFFDF8;padding:2rem;border:1px solid #f0e8d0;border-top:none;">
    <p style="color:#333;line-height:1.6;">You requested a password reset for your account.</p>
    <p style="color:#333;line-height:1.6;">Click the button below to choose a new password.
       This link expires in <strong>1 hour</strong>.</p>
    <div style="text-align:center;margin:2rem 0;">
      <a href="{reset_url}"
         style="background:#8B0000;color:#fff;padding:0.85rem 2rem;border-radius:8px;
                text-decoration:none;font-weight:700;font-size:1rem;display:inline-block;">
        Reset My Password →
      </a>
    </div>
    <p style="color:#999;font-size:0.82rem;">If the button doesn't work, copy this link:<br>
       <a href="{reset_url}" style="color:#8B0000;word-break:break-all;">{reset_url}</a></p>
    <hr style="border:none;border-top:1px solid #eee;margin:1.5rem 0;">
    <p style="color:#bbb;font-size:0.78rem;">If you didn't request this, ignore this email.
       Your password won't change.</p>
  </div>
</div>"""
    msg.attach(MIMEText(text, 'plain'))
    msg.attach(MIMEText(html,  'html'))
    with smtplib.SMTP_SSL('smtp.gmail.com', 465) as srv:
        srv.login(GMAIL_ADDRESS, GMAIL_APP_PASS)
        srv.sendmail(GMAIL_ADDRESS, to_addr, msg.as_string())

def load_pending():
    if not os.path.exists(PENDING_FILE):
        return []
    with open(PENDING_FILE) as f:
        return json.load(f)

def save_pending(pending):
    with open(PENDING_FILE, 'w') as f:
        json.dump(pending, f, indent=2)

def load_inventory():
    if not os.path.exists(INVENTORY_FILE):
        return []
    with open(INVENTORY_FILE, newline='', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        products = list(reader)
    for p in products:
        imgs = [i.strip() for i in p.get('Images','').split(',') if i.strip()]
        p['image_list'] = imgs
        p['valid_images'] = [i for i in imgs if os.path.exists(os.path.join(UPLOAD_FOLDER, i))]
    return products

def save_inventory(products):
    fieldnames = ['SKU','Title','Description','Category','Condition','Price',
                  'Cost Paid','Status','Date Added','Images','Section','Shelf']
    _backup_inventory()
    with open(INVENTORY_FILE, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction='ignore')
        writer.writeheader()
        writer.writerows(products)

def _backup_inventory():
    if not os.path.exists(INVENTORY_FILE):
        return
    ts = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
    dst = os.path.join(BACKUP_FOLDER, f'inventory_{ts}.csv')
    shutil.copy2(INVENTORY_FILE, dst)
    backups = sorted(
        [f for f in os.listdir(BACKUP_FOLDER) if f.endswith('.csv')],
        reverse=True
    )
    for old in backups[MAX_BACKUPS:]:
        os.remove(os.path.join(BACKUP_FOLDER, old))

def get_stats():
    products = load_inventory()
    pending  = load_pending()
    total_value = sum(float(p.get('Price') or 0) for p in products)
    return {
        'total':         len(products),
        'available':     sum(1 for p in products if p.get('Status') == 'Available'),
        'sold':          sum(1 for p in products if p.get('Status') == 'Sold'),
        'reserved':      sum(1 for p in products if p.get('Status') == 'Reserved'),
        'total_value':   total_value,
        'pending_users': len(pending),
    }

# ── Role Definitions ──────────────────────────────────────────────────────────
ROLES = {
    'owner':   {'label': 'Owner',   'color': '#8B0000', 'level': 4},
    'manager': {'label': 'Manager', 'color': '#1a6e4a', 'level': 3},
    'staff':   {'label': 'Staff',   'color': '#1a5a8e', 'level': 2},
    'viewer':  {'label': 'Viewer',  'color': '#666',    'level': 1},
}

PERMISSIONS = {
    'add_product':    ['owner', 'manager', 'staff'],
    'edit_product':   ['owner', 'manager', 'staff'],
    'delete_product': ['owner', 'manager'],
    'price_tag':      ['owner', 'manager', 'staff'],
    'generate_ads':   ['owner', 'manager'],
    'delete_ads':     ['owner', 'manager'],
    'music_upload':   ['owner'],
    'export_import':  ['owner', 'manager'],
    'admin_panel':    ['owner'],
    'view_products':  ['owner', 'manager', 'staff', 'viewer'],
}

def get_user_role(username=None):
    """Return role string for a username (or current session user)."""
    if username is None:
        username = session.get('username', '')
    if username == ADMIN_USER:
        return 'owner'
    users = load_users()
    return users.get(username, {}).get('role', 'viewer')

def can(permission, username=None):
    """Check if a user has a given permission."""
    role = get_user_role(username)
    return role in PERMISSIONS.get(permission, [])

def login_required(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('logged_in'):
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if not can('admin_panel'):
            flash('Owner access required.', 'error')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated

def role_required(permission):
    """Decorator factory: @role_required('generate_ads')"""
    def decorator(f):
        from functools import wraps
        @wraps(f)
        def decorated(*args, **kwargs):
            if not can(permission):
                flash(f'You do not have permission to do that.', 'error')
                return redirect(url_for('dashboard'))
            return f(*args, **kwargs)
        return decorated
    return decorator

def ctx():
    role = get_user_role()
    return dict(
        store_name=STORE_NAME, demo_mode=DEMO_MODE,
        demo_contact_email=CONTACT_EMAIL, stats=get_stats(),
        demo_username=ADMIN_USER, demo_password=ADMIN_PASS,
        is_admin=can('admin_panel'),
        user_role=role,
        user_role_label=ROLES.get(role, {}).get('label', role.title()),
        can_add=can('add_product'),
        can_edit=can('edit_product'),
        can_delete=can('delete_product'),
        can_price_tag=can('price_tag'),
        can_ads=can('generate_ads'),
        can_delete_ads=can('delete_ads'),
        can_export=can('export_import'),
    )

# ── Auth Routes ───────────────────────────────────────────────────────────────
@app.route('/login', methods=['GET','POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username','').strip()
        password = request.form.get('password','')
        # Admin check
        if username == ADMIN_USER and password == ADMIN_PASS:
            session['logged_in'] = True
            session['username']  = ADMIN_USER
            session['is_guest']  = False
            session.permanent    = True
            app.permanent_session_lifetime = datetime.timedelta(hours=8)
            flash('Welcome back, Admin!', 'success')
            return redirect(url_for('dashboard'))
        # Regular user check
        users = load_users()
        if username in users and users[username]['password'] == hash_password(password):
            session['logged_in'] = True
            session['username']  = username
            session['is_guest']  = False
            session.permanent    = True
            app.permanent_session_lifetime = datetime.timedelta(hours=8)
            flash(f'Welcome, {username}!', 'success')
            return redirect(url_for('dashboard'))
        flash('Invalid username or password.', 'error')
    return render_template('login.html', **ctx())

@app.route('/logout')
def logout():
    session.clear()
    flash('You have been logged out.', 'success')
    return redirect(url_for('login'))

@app.route('/guest')
def guest():
    session['logged_in'] = True
    session['username']  = 'guest'
    session['is_guest']  = True
    return redirect(url_for('dashboard'))

@app.route('/signup', methods=['GET','POST'])
def signup():
    if request.method == 'POST':
        username = request.form.get('username','').strip()
        email    = request.form.get('email','').strip()
        password = request.form.get('password','')
        if not username or not password:
            flash('Username and password are required.', 'error')
        elif username == ADMIN_USER:
            flash('That username is reserved.', 'error')
        else:
            users   = load_users()
            pending = load_pending()
            if username in users or any(p['username'] == username for p in pending):
                flash('Username already exists or is pending.', 'error')
            else:
                pending.append({
                    'username':  username,
                    'email':     email,
                    'password':  hash_password(password),
                    'requested': datetime.date.today().isoformat()
                })
                save_pending(pending)
                flash('Account request submitted! Wait for admin approval.', 'success')
                return redirect(url_for('login'))
    return render_template('signup.html', **ctx())

# ── Dashboard ─────────────────────────────────────────────────────────────────
@app.route('/')
@login_required
def dashboard():
    products = load_inventory()
    return render_template('dashboard.html', products=products, **ctx())

# ── Products ──────────────────────────────────────────────────────────────────
@app.route('/product/<sku>')
@login_required
def view_product(sku):
    products = load_inventory()
    product  = next((p for p in products if p['SKU'] == sku), None)
    if not product:
        flash('Product not found.', 'error')
        return redirect(url_for('dashboard'))
    return render_template('product.html', product=product, **ctx())

@app.route('/new', methods=['GET','POST'])
@login_required
@role_required('add_product')
def new_product():
    if session.get('is_guest'):
        flash('Guests cannot add products.', 'error')
        return redirect(url_for('dashboard'))
    if request.method == 'POST':
        sku = request.form.get('sku','').strip().upper()
        if not sku:
            flash('SKU is required.', 'error')
            return render_template('edit_with_ai.html', product={},
                                   categories=CATEGORIES, conditions=CONDITIONS,
                                   statuses=STATUSES, **ctx())
        products = load_inventory()
        if any(p['SKU'] == sku for p in products):
            flash('SKU already exists.', 'error')
            return render_template('edit_with_ai.html', product={},
                                   categories=CATEGORIES, conditions=CONDITIONS,
                                   statuses=STATUSES, **ctx())
        # Handle images
        images = []
        for file in request.files.getlist('images'):
            if file and allowed_file(file.filename):
                ext      = file.filename.rsplit('.', 1)[1].lower()
                filename = f"{sku}_{uuid.uuid4().hex[:8]}.{ext}"
                file.save(os.path.join(UPLOAD_FOLDER, filename))
                images.append(filename)
        product = {
            'SKU':        sku,
            'Title':      request.form.get('title','').strip(),
            'Description':request.form.get('description','').strip(),
            'Category':   request.form.get('category','').strip(),
            'Condition':  request.form.get('condition','Good'),
            'Price':      request.form.get('price','0'),
            'Cost Paid':  request.form.get('cost_paid','') if session.get('username') == ADMIN_USER else '',
            'Status':     request.form.get('status','Available'),
            'Date Added': datetime.date.today().isoformat(),
            'Images':     ','.join(images),
            'Section':    request.form.get('section','').strip(),
            'Shelf':      request.form.get('shelf','').strip(),
        }
        products.append(product)
        save_inventory(products)
        flash(f'Product {sku} created!', 'success')
        return redirect(url_for('view_product', sku=sku))
    return render_template('edit_with_ai.html', product={},
                           categories=CATEGORIES, conditions=CONDITIONS,
                           statuses=STATUSES, **ctx())

@app.route('/edit/<sku>', methods=['GET','POST'])
@login_required
def edit_product(sku):
    if session.get('is_guest'):
        flash('Guests cannot edit products.', 'error')
        return redirect(url_for('dashboard'))
    products = load_inventory()
    idx      = next((i for i, p in enumerate(products) if p['SKU'] == sku), None)
    if idx is None:
        flash('Product not found.', 'error')
        return redirect(url_for('dashboard'))
    product = products[idx]
    if request.method == 'POST':
        for file in request.files.getlist('images'):
            if file and allowed_file(file.filename):
                ext      = file.filename.rsplit('.', 1)[1].lower()
                filename = f"{sku}_{uuid.uuid4().hex[:8]}.{ext}"
                file.save(os.path.join(UPLOAD_FOLDER, filename))
                existing = [i.strip() for i in product.get('Images','').split(',') if i.strip()]
                existing.append(filename)
                product['Images'] = ','.join(existing)
        product['Title']       = request.form.get('title', product['Title']).strip()
        product['Description'] = request.form.get('description', product.get('Description','')).strip()
        product['Category']    = request.form.get('category', product.get('Category','')).strip()
        product['Condition']   = request.form.get('condition', product.get('Condition','Good'))
        product['Price']       = request.form.get('price', product.get('Price','0'))
        product['Status']      = request.form.get('status', product.get('Status','Available'))
        product['Section']     = request.form.get('section', product.get('Section','')).strip()
        product['Shelf']       = request.form.get('shelf', product.get('Shelf','')).strip()
        if session.get('username') == ADMIN_USER:
            product['Cost Paid'] = request.form.get('cost_paid', product.get('Cost Paid',''))
        products[idx] = product
        save_inventory(products)
        flash('Product updated!', 'success')
        return redirect(url_for('view_product', sku=sku))
    return render_template('edit_with_ai.html', product=product,
                           categories=CATEGORIES, conditions=CONDITIONS,
                           statuses=STATUSES, **ctx())

@app.route('/delete/<sku>', methods=['POST'])
@login_required
@role_required('delete_product')
def delete_product(sku):
    if session.get('is_guest'):
        flash('Guests cannot delete products.', 'error')
        return redirect(url_for('dashboard'))
    products = load_inventory()
    products = [p for p in products if p['SKU'] != sku]
    save_inventory(products)
    flash('Product deleted.', 'success')
    return redirect(url_for('dashboard'))

# ── Images ────────────────────────────────────────────────────────────────────
@app.route('/uploads/<filename>')
def serve_upload(filename):
    return send_from_directory(UPLOAD_FOLDER, filename)

@app.route('/delete-image/<sku>', methods=['POST'])
@login_required
def delete_image(sku):
    filename = request.form.get('filename')
    products = load_inventory()
    idx      = next((i for i, p in enumerate(products) if p['SKU'] == sku), None)
    if idx is not None and filename:
        imgs = [i.strip() for i in products[idx].get('Images','').split(',') if i.strip()]
        if filename in imgs:
            imgs.remove(filename)
            products[idx]['Images'] = ','.join(imgs)
            save_inventory(products)
            filepath = os.path.join(UPLOAD_FOLDER, filename)
            if os.path.exists(filepath):
                os.remove(filepath)
    return redirect(url_for('edit_product', sku=sku))

@app.route('/edit-image/<sku>')
@login_required
def edit_image(sku):
    products = load_inventory()
    product  = next((p for p in products if p['SKU'] == sku), None)
    if not product:
        flash('Product not found.', 'error')
        return redirect(url_for('dashboard'))
    return render_template('image_editor.html', product=product, **ctx())

@app.route('/save-image/<sku>', methods=['POST'])
@login_required
def save_image(sku):
    data      = request.json
    image_data= data.get('image_data','')
    filename  = data.get('filename','')
    if image_data and filename:
        header, encoded = image_data.split(',', 1)
        img_bytes = base64.b64decode(encoded)
        filepath  = os.path.join(UPLOAD_FOLDER, filename)
        with open(filepath, 'wb') as f:
            f.write(img_bytes)
        return jsonify({'success': True})
    return jsonify({'success': False})

# ── AI Analysis ───────────────────────────────────────────────────────────────
@app.route('/ai-analyze', methods=['POST'])
@login_required
def ai_analyze():
    api_key = os.environ.get('ANTHROPIC_API_KEY')
    if not api_key:
        return jsonify({'error': 'AI feature not configured. ANTHROPIC_API_KEY missing.'})
    file = request.files.get('image')
    if not file:
        return jsonify({'error': 'No image provided.'})
    img_bytes = file.read()

    # Re-encode via Pillow to ensure valid JPEG — fixes phone HEIC/HEIF and large images
    try:
        from PIL import Image as _Img
        import io as _io
        _pil = _Img.open(_io.BytesIO(img_bytes))
        # Auto-rotate based on EXIF
        try:
            from PIL import ExifTags as _ET
            exif = _pil._getexif()
            if exif:
                orient_key = next((k for k, v in _ET.TAGS.items() if v == 'Orientation'), None)
                if orient_key and orient_key in exif:
                    rot = {3:180, 6:270, 8:90}.get(exif[orient_key])
                    if rot:
                        _pil = _pil.rotate(rot, expand=True)
        except Exception:
            pass
        _pil = _pil.convert('RGB')
        # Resize if too large (phones take huge photos)
        if max(_pil.size) > 1600:
            _pil.thumbnail((1600, 1600), _Img.LANCZOS)
        buf = _io.BytesIO()
        _pil.save(buf, format='JPEG', quality=85)
        img_bytes = buf.getvalue()
    except Exception:
        pass  # If Pillow fails just use original bytes

    img_b64      = base64.b64encode(img_bytes).decode('utf-8')
    content_type = 'image/jpeg'
    try:
        import urllib.request as ur
        import json as _json
        payload = {
            'model': 'claude-haiku-4-5-20251001',
            'max_tokens': 1024,
            'messages': [{
                'role': 'user',
                'content': [
                    {'type': 'image', 'source': {'type': 'base64', 'media_type': content_type, 'data': img_b64}},
                    {'type': 'text', 'text': (
                        'Analyze this thrift store item photo. Respond ONLY with valid JSON:\n'
                        '{"title":"short product name","category":"one of: Furniture/Electronics/Clothing/'
                        'Jewelry/Home Decor/Books/Kitchen/Toys/Tools/Collectibles/Art/Miscellaneous",'
                        '"condition":"one of: New/Like New/Good/Fair/Poor",'
                        '"description":"2-3 sentence description",'
                        '"suggested_price":numeric_value}'
                    )}
                ]
            }]
        }
        req = ur.Request(
            'https://api.anthropic.com/v1/messages',
            data=_json.dumps(payload).encode(),
            headers={
                'x-api-key': api_key,
                'anthropic-version': '2023-06-01',
                'content-type': 'application/json'
            }
        )
        with ur.urlopen(req, timeout=30) as resp:
            result = _json.loads(resp.read())
        text = result['content'][0]['text'].strip()
        # Strip markdown fences if present
        if text.startswith('```'):
            text = text.split('\n', 1)[1].rsplit('```', 1)[0].strip()
        return jsonify(_json.loads(text))
    except Exception as e:
        return jsonify({'error': str(e)})

# ── Price Tag ─────────────────────────────────────────────────────────────────
@app.route('/price-tag/<sku>')
@login_required
@role_required('price_tag')
def price_tag(sku):
    products = load_inventory()
    product  = next((p for p in products if p['SKU'] == sku), None)
    if not product:
        flash('Product not found.', 'error')
        return redirect(url_for('dashboard'))
    return render_template('price_tag.html', product=product, **ctx())

# ── Ad Generator ──────────────────────────────────────────────────────────────
@app.route('/ads')
@login_required
@role_required('generate_ads')
def ad_generator():
    products = load_inventory()
    return render_template('ad_generator.html', products=products, **ctx())

@app.route('/generate-ads', methods=['POST'])
@login_required
@role_required('generate_ads')
def generate_ads():
    # Accept both JSON (from ad_generator.html JS) and form POST
    if request.is_json:
        data          = request.get_json()
        product_list  = data.get('products', [])
        color_theme   = data.get('style', 'red_gold')
        all_products  = load_inventory()
        sku_map       = {p['SKU']: p for p in all_products}
        selected = []
        for item in product_list[:10]:
            sku = item.get('sku','')
            if sku in sku_map:
                selected.append(sku_map[sku])
        use_json_response = True
    else:
        selected_skus = request.form.getlist('selected_products')
        color_theme   = request.form.get('color_theme', 'red_gold')
        products      = load_inventory()
        selected      = [p for p in products if p['SKU'] in selected_skus][:10]
        use_json_response = False
    generated     = []

    themes = {
        'red_gold':      ((139,0,0),    (255,215,0),   (180,20,20)),
        'orange_yellow': ((200,80,0),   (255,237,0),   (230,110,0)),
        'navy_white':    ((0,31,91),    (255,255,255), (0,60,140)),
        'brown_gold':    ((74,44,10),   (201,168,76),  (100,60,20)),
    }
    bg_color, accent_color, header_color = themes.get(color_theme, themes['red_gold'])

    try:
        from PIL import Image, ImageDraw, ImageFont
        use_pillow = True
    except ImportError:
        use_pillow = False

    for p in selected:
        if use_pillow:
            filename = f"ad_{p['SKU']}_{uuid.uuid4().hex[:6]}.jpg"
            filepath = os.path.join(ADS_FOLDER, filename)

            W, H = 800, 600
            img  = Image.new('RGB', (W, H), bg_color)
            draw = ImageDraw.Draw(img)

            # Header bar
            draw.rectangle([0, 0, W, 90], fill=header_color)
            # Footer bar
            draw.rectangle([0, H-60, W, H], fill=header_color)

            # Try to load a font, fall back to default
            try:
                font_lg = ImageFont.truetype('/usr/share/fonts/truetype/dejavu/DejaVuSerif-Bold.ttf', 36)
                font_md = ImageFont.truetype('/usr/share/fonts/truetype/dejavu/DejaVuSerif-Bold.ttf', 28)
                font_sm = ImageFont.truetype('/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf', 20)
                font_xs = ImageFont.truetype('/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf', 16)
            except:
                font_lg = ImageFont.load_default()
                font_md = font_lg
                font_sm = font_lg
                font_xs = font_lg

            # Store name in header
            store_text = STORE_NAME
            bbox = draw.textbbox((0,0), store_text, font=font_lg)
            tw = bbox[2] - bbox[0]
            draw.text(((W-tw)//2, 22), store_text, fill=accent_color, font=font_lg)

            # ── AI-generated headline & tagline ─────────────────────
            ai_headline = p.get('Title', '')[:40]
            ai_tagline  = f"{p.get('Condition','')} · {p.get('Category','')}"
            api_key = os.environ.get('ANTHROPIC_API_KEY')
            if api_key:
                try:
                    import urllib.request as _ur
                    import json as _json
                    _payload = {
                        'model': 'claude-haiku-4-5-20251001',
                        'max_tokens': 120,
                        'messages': [{'role': 'user', 'content':
                            f'''Write a short punchy Facebook Marketplace ad for this thrift store item.
Title: {p.get("Title","")}
Category: {p.get("Category","")}
Condition: {p.get("Condition","")}
Price: ${p.get("Price","0")}
Description: {p.get("Description","")[:200]}

Respond ONLY with JSON: {{"headline":"max 8 words","tagline":"max 12 words"}}'''
                        }]
                    }
                    _req = _ur.Request(
                        'https://api.anthropic.com/v1/messages',
                        data=_json.dumps(_payload).encode(),
                        headers={'x-api-key': api_key, 'anthropic-version': '2023-06-01', 'content-type': 'application/json'}
                    )
                    with _ur.urlopen(_req, timeout=15) as _resp:
                        _result = _json.loads(_resp.read())
                    _text = _result['content'][0]['text'].strip()
                    if _text.startswith('```'): _text = _text.split('```')[0] if '```' in _text else _text
                    _ai = _json.loads(_text)
                    ai_headline = _ai.get('headline', ai_headline)[:40]
                    ai_tagline  = _ai.get('tagline',  ai_tagline)[:50]
                except:
                    pass  # Fall back to product title if AI fails

            # ── Product image (EXIF auto-rotate) ─────────────────────
            img_y_start = 100
            img_area_h  = 270
            if p.get('valid_images'):
                try:
                    from PIL.ExifTags import TAGS
                    prod_img_path = os.path.join(UPLOAD_FOLDER, p['valid_images'][0])
                    prod_img = Image.open(prod_img_path)
                    # Auto-rotate based on EXIF orientation
                    try:
                        exif = prod_img._getexif()
                        if exif:
                            orientation_key = next((k for k,v in TAGS.items() if v == 'Orientation'), None)
                            if orientation_key and orientation_key in exif:
                                orientation = exif[orientation_key]
                                rotations = {3:180, 6:270, 8:90}
                                if orientation in rotations:
                                    prod_img = prod_img.rotate(rotations[orientation], expand=True)
                    except:
                        pass
                    prod_img = prod_img.convert('RGB')
                    prod_img.thumbnail((340, img_area_h))
                    px = (W - prod_img.width) // 2
                    py = img_y_start + (img_area_h - prod_img.height) // 2
                    img.paste(prod_img, (px, py))
                except:
                    pass

            # ── AI Headline ───────────────────────────────────────────
            bbox  = draw.textbbox((0,0), ai_headline, font=font_md)
            tw    = bbox[2] - bbox[0]
            # Word wrap if too wide
            if tw > W - 40:
                words = ai_headline.split()
                mid   = len(words) // 2
                line1 = ' '.join(words[:mid])
                line2 = ' '.join(words[mid:])
                b1 = draw.textbbox((0,0), line1, font=font_md)
                b2 = draw.textbbox((0,0), line2, font=font_md)
                draw.text(((W-(b1[2]-b1[0]))//2, 382), line1, fill=accent_color, font=font_md)
                draw.text(((W-(b2[2]-b2[0]))//2, 414), line2, fill=accent_color, font=font_md)
                price_y = 448
            else:
                draw.text(((W-tw)//2, 390), ai_headline, fill=accent_color, font=font_md)
                price_y = 430

            # ── Price badge ───────────────────────────────────────────
            price_text = f"${p.get('Price','0.00')}"
            badge_w, badge_h = 200, 56
            bx = (W - badge_w) // 2
            draw.rounded_rectangle([bx, price_y, bx+badge_w, price_y+badge_h], radius=12, fill=accent_color)
            bbox  = draw.textbbox((0,0), price_text, font=font_lg)
            tw    = bbox[2] - bbox[0]
            th    = bbox[3] - bbox[1]
            draw.text((bx+(badge_w-tw)//2, price_y+(badge_h-th)//2 - 2), price_text, fill=bg_color, font=font_lg)

            # ── AI Tagline ────────────────────────────────────────────
            bbox   = draw.textbbox((0,0), ai_tagline, font=font_xs)
            tw     = bbox[2] - bbox[0]
            draw.text(((W-tw)//2, price_y + badge_h + 8), ai_tagline, fill=accent_color, font=font_xs)

            # Footer
            product_url = f"https://edclawd.pythonanywhere.com/product/{p['SKU']}"
            footer = f"125 W Swannanoa Ave, Liberty NC 27298  |  View: {product_url}"
            bbox   = draw.textbbox((0,0), footer, font=font_xs)
            tw     = bbox[2] - bbox[0]
            draw.text(((W-tw)//2, H-42), footer, fill=accent_color, font=font_xs)

            img.save(filepath, 'JPEG', quality=90)
        else:
            # HTML fallback
            filename = f"ad_{p['SKU']}_{uuid.uuid4().hex[:6]}.html"
            filepath = os.path.join(ADS_FOLDER, filename)
            bg  = '#{:02x}{:02x}{:02x}'.format(*bg_color)
            acc = '#{:02x}{:02x}{:02x}'.format(*accent_color)
            html = f"""<!DOCTYPE html><html><head><meta charset="UTF-8">
<style>body{{margin:0;font-family:Georgia,serif;background:{bg};color:{acc};text-align:center;padding:20px}}
.store{{font-size:1.4rem;font-weight:bold;padding-bottom:8px;margin-bottom:12px}}
.price{{font-size:2rem;font-weight:bold;border:3px solid {acc};display:inline-block;padding:8px 20px;border-radius:8px;margin:10px 0}}
</style></head><body>
<div class="store">✨ {STORE_NAME} ✨</div>
<div class="title">{p.get('Title','')}</div>
<div class="price">${p.get('Price','0.00')}</div>
<div>📍 Asheboro, NC</div></body></html>"""
            with open(filepath, 'w') as f:
                f.write(html)

        # ── Build HTML wrapper with JPEG + clickable button ──────
        product_url  = f"https://edclawd.pythonanywhere.com/product/{p['SKU']}"
        html_filename = filename.replace('.jpg', '.html') if filename.endswith('.jpg') else filename
        html_filepath = os.path.join(ADS_FOLDER, html_filename)
        bg_hex  = '#{:02x}{:02x}{:02x}'.format(*bg_color)
        acc_hex = '#{:02x}{:02x}{:02x}'.format(*accent_color)
        hdr_hex = '#{:02x}{:02x}{:02x}'.format(*header_color)
        img_tag = f'<img src="/ads/{filename}" style="width:100%;display:block;">' if filename.endswith('.jpg') else ''
        html_content = f"""<!DOCTYPE html>
<html>
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<style>
  body {{ margin:0; padding:0; background:#1a1a1a; display:flex; justify-content:center; align-items:center; min-height:100vh; font-family:Georgia,serif; }}
  .ad-wrap {{ max-width:800px; width:100%; background:{bg_hex}; border-radius:12px; overflow:hidden; box-shadow:0 8px 32px rgba(0,0,0,0.5); }}
  .ad-img {{ width:100%; display:block; }}
  .ad-footer {{ background:{hdr_hex}; padding:1rem 1.5rem; display:flex; justify-content:space-between; align-items:center; flex-wrap:wrap; gap:0.75rem; }}
  .ad-address {{ color:{acc_hex}; font-size:0.85rem; }}
  .view-btn {{
    background:{acc_hex}; color:{bg_hex}; border:none; padding:0.75rem 1.75rem;
    border-radius:8px; font-size:1rem; font-weight:bold; cursor:pointer;
    text-decoration:none; display:inline-block; transition:opacity 0.2s;
  }}
  .view-btn:hover {{ opacity:0.85; }}
</style>
</head>
<body>
  <div class="ad-wrap">
    {img_tag}
    <div class="ad-footer">
      <div class="ad-address">📍 125 W Swannanoa Ave, Liberty NC 27298</div>
      <a href="{product_url}" class="view-btn" target="_blank">🛍️ View Product</a>
    </div>
  </div>
</body>
</html>"""
        with open(html_filepath, 'w') as hf:
            hf.write(html_content)

        generated.append({'filename': html_filename, 'product_title': p.get('Title',''), 'type': 'html'})

    if use_json_response:
        return jsonify({'success': True, 'files': generated})
    return render_template('ads.html', generated=[g['filename'] for g in generated], **ctx())


@app.route('/ads/<filename>')
def view_ad(filename):
    return send_from_directory(ADS_FOLDER, filename)

@app.route('/download-ad/<filename>')
@login_required
def download_ad(filename):
    return send_from_directory(ADS_FOLDER, filename, as_attachment=True)

# ── Music Library ──────────────────────────────────────────────────────────────
ALLOWED_MUSIC = {'mp3', 'wav', 'm4a', 'aac'}

def allowed_music(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_MUSIC

@app.route('/music/list')
@login_required
def music_list():
    files = []
    if os.path.exists(MUSIC_FOLDER):
        for fname in sorted(os.listdir(MUSIC_FOLDER)):
            fpath = os.path.join(MUSIC_FOLDER, fname)
            if not os.path.isfile(fpath):
                continue
            ext = fname.rsplit('.', 1)[-1].lower() if '.' in fname else ''
            if ext not in ALLOWED_MUSIC:
                continue
            size_kb = round(os.path.getsize(fpath) / 1024, 1)
            name = fname.rsplit('.', 1)[0].replace('_', ' ').replace('-', ' ')
            files.append({'filename': fname, 'name': name, 'ext': ext.upper(), 'size_kb': size_kb})
    return jsonify({'success': True, 'files': files})

@app.route('/music/upload', methods=['POST'])
@login_required
def music_upload():
    if session.get('username') != ADMIN_USER:
        return jsonify({'success': False, 'error': 'Admin only'}), 403
    f = request.files.get('music_file')
    if not f or not f.filename:
        return jsonify({'success': False, 'error': 'No file provided'})
    if not allowed_music(f.filename):
        return jsonify({'success': False, 'error': 'Invalid file type. Use MP3, WAV, M4A or AAC.'})
    fname = secure_filename(f.filename)
    os.makedirs(MUSIC_FOLDER, exist_ok=True)
    f.save(os.path.join(MUSIC_FOLDER, fname))
    size_kb = round(os.path.getsize(os.path.join(MUSIC_FOLDER, fname)) / 1024, 1)
    name = fname.rsplit('.', 1)[0].replace('_', ' ').replace('-', ' ')
    ext  = fname.rsplit('.', 1)[-1].upper()
    return jsonify({'success': True, 'filename': fname, 'name': name, 'ext': ext, 'size_kb': size_kb})

@app.route('/music/delete/<filename>', methods=['POST'])
@login_required
def music_delete(filename):
    if session.get('username') != ADMIN_USER:
        return jsonify({'success': False, 'error': 'Admin only'}), 403
    fpath = os.path.join(MUSIC_FOLDER, secure_filename(filename))
    try:
        if os.path.exists(fpath):
            os.remove(fpath)
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/music/<filename>')
def serve_music(filename):
    return send_from_directory(MUSIC_FOLDER, filename)


@app.route('/delete-ad/<filename>', methods=['POST'])
@login_required
@role_required('delete_ads')
def delete_ad(filename):
    filepath = os.path.join(ADS_FOLDER, filename)
    try:
        if os.path.exists(filepath):
            os.remove(filepath)
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/ad-gallery')
@login_required
@role_required('generate_ads')
def ad_gallery():
    ads = []
    if os.path.exists(ADS_FOLDER):
        for fname in sorted(os.listdir(ADS_FOLDER), reverse=True):
            fpath = os.path.join(ADS_FOLDER, fname)
            if not os.path.isfile(fpath):
                continue
            ext = fname.rsplit('.', 1)[-1].lower() if '.' in fname else ''
            if ext not in ('jpg', 'jpeg', 'png', 'mp4', 'html'):
                continue
            stat = os.stat(fpath)
            size_kb = round(stat.st_size / 1024, 1)
            mtime = stat.st_mtime
            from datetime import datetime as _dt
            created = _dt.fromtimestamp(mtime).strftime('%b %d, %Y %I:%M %p')
            # Try to extract product name from filename
            # Typical pattern: adtype_SKU_timestamp.ext or SKU_timestamp.ext
            label = fname.replace('_', ' ').rsplit('.', 1)[0]
            ad_type = 'video' if ext == 'mp4' else ('html' if ext == 'html' else 'image')
            ads.append({
                'filename': fname,
                'label': label,
                'type': ad_type,
                'ext': ext,
                'size_kb': size_kb,
                'created': created,
                'url': f'/ads/{fname}',
                'download_url': f'/download-ad/{fname}',
            })
    return render_template('ad_gallery.html', ads=ads, **ctx())

@app.route('/generate-video-ad', methods=['POST'])
@login_required
@role_required('generate_ads')
def generate_video_ad():
    """Generate an MP4 video ad for a single product."""
    import subprocess, tempfile, wave, struct, math
    try:
        import numpy as np
        HAS_NUMPY = True
    except ImportError:
        HAS_NUMPY = False

    data       = request.get_json() or {}
    sku        = data.get('sku', '')
    style      = data.get('style', 'elegant')
    music_file = data.get('music_file', '')

    # -- Load product --
    products = load_inventory()
    product  = next((p for p in products if p['SKU'] == sku), None)
    if not product:
        return jsonify({'success': False, 'error': f'Product {sku} not found'})

    # -- Find product image --
    image_file   = data.get('image_file', '')
    product_imgs = []
    if os.path.exists(UPLOAD_FOLDER):
        for fn in os.listdir(UPLOAD_FOLDER):
            if fn.startswith(sku + '_') or fn.startswith(sku + '-'):
                product_imgs.append(fn)
        product_imgs.sort()

    chosen_img = None
    if image_file and os.path.exists(os.path.join(UPLOAD_FOLDER, image_file)):
        chosen_img = os.path.join(UPLOAD_FOLDER, image_file)
    elif product_imgs:
        chosen_img = os.path.join(UPLOAD_FOLDER, product_imgs[0])

    # -- Style colours --
    themes = {
        'elegant':  {'bg': (26,10,10),   'accent': (139,0,0),    'text': (255,215,0),   'sub': (230,200,150)},
        'bright':   {'bg': (255,255,255), 'accent': (220,50,50),  'text': (30,30,30),    'sub': (80,80,80)},
        'minimal':  {'bg': (245,245,245), 'accent': (50,50,50),   'text': (20,20,20),    'sub': (100,100,100)},
        'vintage':  {'bg': (245,235,210), 'accent': (120,60,20),  'text': (60,30,10),    'sub': (120,90,60)},
    }
    theme = themes.get(style, themes['elegant'])

    # -- Build frames using PIL --
    try:
        from PIL import Image, ImageDraw, ImageFont, ImageOps, ImageFilter
    except ImportError:
        return jsonify({'success': False, 'error': 'Pillow not available'})

    W, H   = 1080, 1080
    FPS    = 30
    SECS   = 8
    FRAMES = FPS * SECS

    title       = product.get('Title', 'Item')
    price_raw   = product.get('Price', '')
    try:
        price_str = '${:.2f}'.format(float(str(price_raw).replace('$','').strip()))
    except Exception:
        price_str = str(price_raw)
    store_line  = STORE_NAME

    font_bold = font_body = font_price = font_store = None
    for fp in ['/usr/share/fonts/truetype/dejavu/DejaVuSerif-Bold.ttf',
               '/usr/share/fonts/truetype/freefont/FreeSerifBold.ttf']:
        if os.path.exists(fp):
            try:
                font_bold  = ImageFont.truetype(fp, 72)
                font_price = ImageFont.truetype(fp, 96)
                font_store = ImageFont.truetype(fp, 38)
                break
            except Exception:
                pass
    for fp in ['/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
               '/usr/share/fonts/truetype/freefont/FreeSans.ttf']:
        if os.path.exists(fp):
            try:
                font_body = ImageFont.truetype(fp, 48)
                break
            except Exception:
                pass
    if not font_bold:  font_bold  = ImageFont.load_default()
    if not font_body:  font_body  = ImageFont.load_default()
    if not font_price: font_price = font_bold
    if not font_store: font_store = font_body

    # Load & prepare product image
    prod_img = None
    if chosen_img:
        try:
            pi = Image.open(chosen_img)
            pi = ImageOps.exif_transpose(pi)
            pi = pi.convert('RGBA')
            # Fit inside 560x560
            pi.thumbnail((560, 560), Image.LANCZOS)
            prod_img = pi
        except Exception:
            prod_img = None

    tmpdir = tempfile.mkdtemp()
    frame_pattern = os.path.join(tmpdir, 'frame_%05d.png')

    bg_color  = theme['bg']
    acc_color = theme['accent']
    txt_color = theme['text']
    sub_color = theme['sub']

    def wrap_text(text, font, max_w, draw):
        words = text.split()
        lines, cur = [], ''
        for w in words:
            test = (cur + ' ' + w).strip()
            bbox = draw.textbbox((0,0), test, font=font)
            if bbox[2] - bbox[0] <= max_w:
                cur = test
            else:
                if cur: lines.append(cur)
                cur = w
        if cur: lines.append(cur)
        return lines

    for i in range(FRAMES):
        t = i / FRAMES  # 0..1

        img  = Image.new('RGB', (W, H), bg_color)
        draw = ImageDraw.Draw(img)

        # Animated accent bar at top
        bar_w = int(W * min(1.0, t * 3))
        draw.rectangle([0, 0, bar_w, 8], fill=acc_color)
        # Bottom bar
        draw.rectangle([0, H-8, W, H], fill=acc_color)

        # Decorative corner lines
        draw.rectangle([0, 0, 6, H], fill=acc_color)
        draw.rectangle([W-6, 0, W, H], fill=acc_color)

        # Product image with fade-in
        img_alpha = min(1.0, t * 2.5)
        if prod_img:
            pi_copy = prod_img.copy()
            alpha = pi_copy.split()[-1] if pi_copy.mode == 'RGBA' else None
            if alpha:
                alpha = alpha.point(lambda x: int(x * img_alpha))
                pi_copy.putalpha(alpha)
            # Center image in upper portion
            px = (W - pi_copy.width) // 2
            py = 80 + int((1 - img_alpha) * 40)
            img.paste(pi_copy, (px, py), pi_copy)
        else:
            # Placeholder box
            box_alpha = int(255 * img_alpha)
            ph = Image.new('RGBA', (400, 400), acc_color + (box_alpha,))
            img.paste(ph, (340, 100), ph)
            draw.text((540, 300), '🛍', font=font_price, fill=txt_color, anchor='mm')

        # Text slide-up animation
        text_y_offset = int((1 - min(1.0, t * 2)) * 60)

        # Title
        title_y = 670 + text_y_offset
        title_alpha = min(255, int(min(1.0, t * 2.5) * 255))
        lines = wrap_text(title, font_bold, W - 120, draw)
        for li, line in enumerate(lines[:2]):
            draw.text((W//2, title_y + li * 80), line, font=font_bold,
                      fill=txt_color + (title_alpha,) if len(txt_color)==3 else txt_color,
                      anchor='mm')

        # Price with pulse
        pulse = 1.0 + 0.03 * math.sin(t * math.pi * 6)
        price_y = title_y + (len(lines[:2])) * 80 + 30
        price_font = font_price
        draw.text((W//2, price_y), price_str, font=price_font, fill=acc_color, anchor='mm')

        # Store name
        store_alpha = min(255, int(max(0, t - 0.4) * 4 * 255))
        draw.text((W//2, H - 55), store_line, font=font_store,
                  fill=sub_color, anchor='mm')

        img.save(os.path.join(tmpdir, f'frame_{i:05d}.png'))

    # -- Music --
    music_path = None
    if music_file:
        candidate = os.path.join(MUSIC_FOLDER, music_file)
        if os.path.exists(candidate):
            music_path = candidate

    if not music_path:
        # Generate simple numpy/wave tone
        wav_path = os.path.join(tmpdir, 'music.wav')
        sr = 44100
        duration = SECS
        if HAS_NUMPY:
            t_arr = np.linspace(0, duration, int(sr * duration))
            mood = data.get('music', 'calm')
            if mood == 'upbeat':
                freq = [440, 523, 659, 784]
            elif mood == 'dramatic':
                freq = [220, 277, 330, 415]
            else:
                freq = [330, 392, 494, 587]
            audio = np.zeros_like(t_arr)
            for fi, f in enumerate(freq):
                seg = int(len(t_arr) / len(freq))
                start = fi * seg
                end = min(start + seg, len(t_arr))
                env = np.ones(end - start)
                fade = min(int(sr * 0.1), (end - start) // 2)
                env[:fade] = np.linspace(0, 1, fade)
                env[-fade:] = np.linspace(1, 0, fade)
                audio[start:end] += 0.3 * np.sin(2 * np.pi * f * t_arr[start:end]) * env
            audio = np.clip(audio, -1, 1)
            samples = (audio * 32767).astype(np.int16)
            with wave.open(wav_path, 'w') as wf:
                wf.setnchannels(1); wf.setsampwidth(2); wf.setframerate(sr)
                wf.writeframes(samples.tobytes())
        else:
            with wave.open(wav_path, 'w') as wf:
                wf.setnchannels(1); wf.setsampwidth(2); wf.setframerate(44100)
                for _ in range(44100 * SECS):
                    wf.writeframes(struct.pack('<h', 0))
        music_path = wav_path

    # -- Encode with ffmpeg --
    safe_title = ''.join(c if c.isalnum() or c in '-_' else '_' for c in title[:30])
    out_filename = f'video_{sku}_{safe_title}_{uuid.uuid4().hex[:6]}.mp4'
    out_path = os.path.join(ADS_FOLDER, out_filename)
    os.makedirs(ADS_FOLDER, exist_ok=True)

    cmd = [
        'ffmpeg', '-y',
        '-framerate', str(FPS),
        '-i', frame_pattern,
        '-i', music_path,
        '-c:v', 'libx264', '-preset', 'fast', '-crf', '23',
        '-pix_fmt', 'yuv420p',
        '-c:a', 'aac', '-b:a', '128k',
        '-shortest',
        '-t', str(SECS),
        out_path
    ]
    result = subprocess.run(cmd, capture_output=True, text=True, timeout=120)

    # Cleanup temp frames
    try:
        import shutil as _sh
        _sh.rmtree(tmpdir, ignore_errors=True)
    except Exception:
        pass

    if result.returncode != 0:
        return jsonify({'success': False, 'error': 'ffmpeg failed: ' + result.stderr[-500:]})

    return jsonify({
        'success': True,
        'filename': out_filename,
        'product_title': title,
        'url': f'/ads/{out_filename}',
    })

# ── Listing Generator ───────────────────────────────────────────────────────
@app.route('/listing-generator')
@login_required
def listing_generator():
    products = load_inventory()
    return render_template('listing_generator.html', products=products, **ctx())

@app.route('/generate-listing', methods=['POST'])
@login_required
def generate_listing():
    data     = request.get_json()
    product  = data.get('product', {})
    platform = data.get('platform', 'facebook')
    api_key  = os.environ.get('ANTHROPIC_API_KEY')

    title     = product.get('title', '')
    price     = product.get('price', '0')
    category  = product.get('category', '')
    condition = product.get('condition', '')
    desc      = product.get('description', '')
    sku       = product.get('sku', '')
    product_url = 'https://edclawd.pythonanywhere.com/product/' + sku
    store_info  = 'Liberty Emporium & Thrift\n125 W Swannanoa Ave, Liberty NC 27298\nView item: ' + product_url

    platform_prompts = {
        'facebook': (
            'Write a Facebook Marketplace listing for this thrift store item.\n'
            'Title: ' + title + ' | Price: $' + price + ' | Category: ' + category + ' | Condition: ' + condition + '\n'
            'Description: ' + desc + '\n'
            'Store: Liberty Emporium & Thrift, 125 W Swannanoa Ave, Liberty NC 27298\n'
            'View online: ' + product_url + '\n\n'
            'Respond ONLY with JSON (no markdown):\n'
            '{"title":"catchy listing title max 60 chars","price":"$' + price + '","condition":"' + condition + '","description":"3-4 sentences friendly tone end with store address and product link","location":"Liberty, NC 27298"}'
        ),
        'craigslist': (
            'Write a Craigslist for-sale listing for this thrift store item.\n'
            'Title: ' + title + ' | Price: $' + price + ' | Category: ' + category + ' | Condition: ' + condition + '\n'
            'Description: ' + desc + '\n'
            'Store: Liberty Emporium & Thrift, 125 W Swannanoa Ave, Liberty NC 27298\n'
            'View online: ' + product_url + '\n\n'
            'Respond ONLY with JSON (no markdown):\n'
            '{"title":"listing title max 70 chars","price":"$' + price + '","condition":"","description":"3-5 sentences practical tone include address and link","location":"Liberty, NC 27298"}'
        ),
        'instagram': (
            'Write an Instagram caption for selling this thrift store item.\n'
            'Title: ' + title + ' | Price: $' + price + ' | Category: ' + category + ' | Condition: ' + condition + '\n'
            'Description: ' + desc + '\n'
            'Store: Liberty Emporium & Thrift, 125 W Swannanoa Ave, Liberty NC 27298\n'
            'View online: ' + product_url + '\n\n'
            'Respond ONLY with JSON (no markdown):\n'
            '{"title":"","price":"$' + price + '","condition":"","description":"Engaging caption with emojis price condition store address product link and 5 hashtags","location":""}'
        ),
    }

    prompt = platform_prompts.get(platform, platform_prompts['facebook'])

    if not api_key:
        fallback_desc = desc + '\n\n' + store_info
        return jsonify({'title': title, 'price': '$' + price, 'condition': condition,
                        'description': fallback_desc, 'location': 'Liberty, NC 27298'})
    try:
        import urllib.request as _ur
        import json as _json
        payload = {
            'model': 'claude-haiku-4-5-20251001',
            'max_tokens': 500,
            'messages': [{'role': 'user', 'content': prompt}]
        }
        req = _ur.Request(
            'https://api.anthropic.com/v1/messages',
            data=_json.dumps(payload).encode(),
            headers={'x-api-key': api_key, 'anthropic-version': '2023-06-01', 'content-type': 'application/json'}
        )
        with _ur.urlopen(req, timeout=30) as resp:
            result = _json.loads(resp.read())
        text = result['content'][0]['text'].strip()
        if text.startswith('```'):
            text = text.split('\n', 1)[1].rsplit('```', 1)[0].strip()
        return jsonify(_json.loads(text))
    except Exception as e:
        return jsonify({'error': str(e)})

# ── Jay Resume / About Page ──────────────────────────────────────────────────
@app.route('/contact')
def contact():
    return render_template('jay_resume.html')

# ── Seasonal Sale (stub) ──────────────────────────────────────────────────────
SALE_FILE = os.path.join(BASE_DIR, 'sale_state.json')

def load_sale():
    if not os.path.exists(SALE_FILE):
        return {'active': False}
    with open(SALE_FILE) as f:
        return json.load(f)

@app.route('/seasonal-sale', methods=['GET','POST'])
@login_required
@admin_required
def seasonal_sale():
    sale_state = load_sale()
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'activate':
            sale_state = {
                'active': True,
                'category': request.form.get('category',''),
                'discount_percent': int(request.form.get('discount_percent', 10))
            }
        else:
            sale_state = {'active': False}
        with open(SALE_FILE, 'w') as f:
            json.dump(sale_state, f)
        flash('Sale settings updated!', 'success')
    return render_template('seasonal_sale.html', sale_state=sale_state,
                           categories=CATEGORIES, **ctx())

# ── Export ────────────────────────────────────────────────────────────────────
@app.route('/export')
@login_required
def export_inventory():
    return send_file(INVENTORY_FILE, as_attachment=True, download_name='inventory.csv')

@app.route('/export-square')
@login_required
@role_required('export_import')
def export_square():
    """Export inventory as a Square-compatible XLSX matching the official template exactly."""
    import io, re as _re, tempfile
    try:
        from openpyxl import load_workbook
    except ImportError:
        return 'openpyxl not available', 500

    BASE_URL  = 'https://edclawd.pythonanywhere.com'
    TMPL_PATH = os.path.join(BASE_DIR, 'square_template.xlsx')

    # If a saved template exists use it, otherwise fall back to a blank one
    if os.path.exists(TMPL_PATH):
        wb = load_workbook(TMPL_PATH)
        ws = wb['Items']
        # Clear old data rows (keep rows 1-6: instructions + header)
        for row in ws.iter_rows(min_row=7):
            for cell in row:
                cell.value = None
    else:
        # Build minimal structure matching Square format
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = Workbook()
        ws = wb.active
        ws.title = 'Items'
        # Row 2: Instructions header
        ws['A2'] = 'Instructions'
        ws['A2'].font = Font(bold=True)
        ws.merge_cells('A2:D2')
        # Row 3: Instructions text
        ws['A3'] = ('Fill in your catalog information below. Avoid common errors: '
                    '1) each row contains an item name; 2) GTINs are 8,12,13, or 14 digits.')
        ws['A3'].alignment = Alignment(wrap_text=True)
        ws.merge_cells('A3:J3')
        ws.row_dimensions[3].height = 71
        # Row 5: example hint row
        ex = ['Custom or created by Square','Created by Square','Example item',
              'Example item name for customers to see','Small','A12345',
              '100% organic soft cotton t-shirt','','','Search engine display name',
              'Search engine description','','653412483659','Visible, Hidden, Unavailable',
              'Physical, Prepared food and beverage, Digital, Donation, Event, Membership, Other',
              '','Social media link title','Social media link description',
              '','','','','24.99','20.99','Y or N','Y or N','Y or N','Y or N','Y or N',
              'Sizes','Small','10','15','Y or N','5']
        for ci, v in enumerate(ex, start=1):
            ws.cell(row=5, column=ci, value=v)
        # Row 6: column headers
        HEADERS = [
            'Reference Handle','Token','Item Name','Customer-facing Name',
            'Variation Name','SKU','Description','Categories',
            'Reporting Category','SEO Title','SEO Description','Permalink',
            'GTIN','Square Online Item Visibility','Item Type','Weight (lb)',
            'Social Media Link Title','Social Media Link Description',
            'Shipping Enabled','Self-serve Ordering Enabled','Delivery Enabled',
            'Pickup Enabled','Price','Online Sale Price','Archived',
            'Sellable','Contains Alcohol','Stockable','Skip Detail Screen in POS',
            'Option Name 1','Option Value 1',
            'Current Quantity Liberty Emporium','New Quantity Liberty Emporium',
            'Stock Alert Enabled Liberty Emporium','Stock Alert Count Liberty Emporium',
        ]
        hdr_fill = PatternFill('solid', fgColor='E7F0FE')
        for ci, h in enumerate(HEADERS, start=1):
            cell = ws.cell(row=6, column=ci, value=h)
            cell.font = Font(bold=True)
            cell.fill = hdr_fill
        for col in ws.column_dimensions:
            ws.column_dimensions[col].width = 9.5

    # Write product data starting at row 7
    products = load_inventory()
    data_row  = 7
    for p in products:
        title  = p.get('Title', '')
        sku    = p.get('SKU', '')
        status = p.get('Status', 'Available')

        slug       = _re.sub(r'[^a-z0-9]+', '-', title.lower()).strip('-')
        ref_handle = '#' + slug + '-regular'

        price_raw = p.get('Price', '')
        try:
            price = '{:.2f}'.format(float(str(price_raw).replace('$', '').strip()))
        except Exception:
            price = ''

        archived = 'Y' if status.lower() in ('sold', 'archived', 'inactive') else 'N'

        raw_imgs  = [i.strip() for i in p.get('Images', '').split(',') if i.strip()]
        valid_imgs = [i for i in raw_imgs if os.path.exists(os.path.join(UPLOAD_FOLDER, i))]

        row_data = [
            ref_handle, '', title, '', 'Regular', sku,
            p.get('Description', ''), p.get('Category', ''),
            '', '', '', '', '', 'unavailable', 'Physical good', '',
            '', '', 'Y', '', '', 'Y',
            price, '', archived, '', '', '', '',
            '', '', '0', '', '', '',
        ]
        for col_idx, val in enumerate(row_data, start=1):
            ws.cell(row=data_row, column=col_idx, value=val)

        # Add image URLs as a note in a column after AI (col 35+)
        for img_idx, img in enumerate(valid_imgs[:4], start=36):
            ws.cell(row=data_row, column=img_idx,
                    value=BASE_URL + '/uploads/' + img)

        data_row += 1

    # Add image URL headers if any products have images
    if any(p.get('Images') for p in products):
        for i, lbl in enumerate(['Image URL 1','Image URL 2','Image URL 3','Image URL 4'], start=36):
            ws.cell(row=6, column=i, value=lbl)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return app.response_class(
        buf.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename=square_import.xlsx'}
    )

@app.route('/import-square', methods=['GET', 'POST'])
@login_required
@role_required('export_import')
def import_square():
    if request.method == 'POST':
        f = request.files.get('square_csv')
        if not f:
            flash('No file uploaded.', 'error')
            return redirect(url_for('import_square'))
        import io, csv as _csv
        stream = io.StringIO(f.stream.read().decode('utf-8-sig'))
        reader = _csv.DictReader(stream)
        products = load_inventory()
        sku_map = {p['SKU']: p for p in products}
        added = updated = 0
        for row in reader:
            sku = (row.get('SKU') or row.get('Token') or '').strip()
            name = (row.get('Item Name') or row.get('Name') or '').strip()
            if not name:
                continue
            price_raw = row.get('Price', '0')
            try:
                price = float(str(price_raw).replace('$', '').replace(',', '').strip())
            except Exception:
                price = 0.0
            if sku and sku in sku_map:
                sku_map[sku]['Title'] = name
                sku_map[sku]['Price'] = price
                sku_map[sku]['Category'] = row.get('Category', sku_map[sku].get('Category', ''))
                updated += 1
            else:
                if not sku:
                    sku = 'SQ-' + str(uuid.uuid4())[:8].upper()
                new_p = {
                    'SKU': sku, 'Title': name,
                    'Description': row.get('Description', ''),
                    'Category': row.get('Category', ''),
                    'Condition': 'Good', 'Price': price,
                    'Cost Paid': 0.0, 'Status': 'Available',
                    'Date Added': datetime.date.today().isoformat(),
                    'Images': '', 'Section': '', 'Shelf': ''
                }
                products.append(new_p)
                added += 1
        save_inventory(products)
        flash(f'Square import complete: {added} added, {updated} updated.', 'success')
        return redirect(url_for('dashboard'))
    return render_template('import_square.html', **ctx())

# ── Admin – Users ─────────────────────────────────────────────────────────────
@app.route('/admin/users')
@login_required
@admin_required
def admin_users():
    users   = load_users()
    pending = load_pending()
    approved = [
        {
            'username': u,
            'email': data.get('email',''),
            'role': data.get('role','viewer'),
            'role_label': ROLES.get(data.get('role','viewer'), {}).get('label','Viewer'),
            'role_color': ROLES.get(data.get('role','viewer'), {}).get('color','#666'),
            'approved': data.get('approved',''),
        }
        for u, data in users.items()
    ]
    return render_template('admin_users.html', users=users, pending=pending, approved=approved, roles=ROLES, **ctx())

@app.route('/admin/approve/<username>', methods=['POST'])
@login_required
@admin_required
def approve_user(username):
    pending = load_pending()
    user    = next((p for p in pending if p['username'] == username), None)
    if user:
        role = request.form.get('role', 'staff')
        if role not in ROLES:
            role = 'staff'
        users = load_users()
        users[username] = {
            'password': user['password'],
            'email': user.get('email', ''),
            'role': role,
            'approved': datetime.date.today().isoformat(),
        }
        save_users(users)
        pending = [p for p in pending if p['username'] != username]
        save_pending(pending)
        flash(f'{username} approved as {ROLES[role]["label"]}!', 'success')
    return redirect(url_for('admin_users'))

@app.route('/admin/change-role/<username>', methods=['POST'])
@login_required
@admin_required
def change_role(username):
    role = request.form.get('role', 'staff')
    if role not in ROLES or username == ADMIN_USER:
        flash('Invalid role.', 'error')
        return redirect(url_for('admin_users'))
    users = load_users()
    if username in users:
        users[username]['role'] = role
        save_users(users)
        flash(f'{username} is now a {ROLES[role]["label"]}.', 'success')
    return redirect(url_for('admin_users'))

@app.route('/admin/reject/<username>', methods=['POST'])
@login_required
@admin_required
def reject_user(username):
    pending = [p for p in load_pending() if p['username'] != username]
    save_pending(pending)
    flash(f'User {username} rejected.', 'success')
    return redirect(url_for('admin_users'))

@app.route('/admin/remove/<username>', methods=['POST'])
@login_required
@admin_required
def remove_user(username):
    users = load_users()
    users.pop(username, None)
    save_users(users)
    flash(f'User {username} removed.', 'success')
    return redirect(url_for('admin_users'))

@app.route('/admin/reset-password/<username>', methods=['POST'])
@login_required
@admin_required
def admin_reset_password(username):
    """Admin force-sets a user password."""
    if username == ADMIN_USER:
        flash('Cannot reset the owner password here.', 'error')
        return redirect(url_for('admin_users'))
    new_pw = request.form.get('new_password', '').strip()
    if len(new_pw) < 6:
        flash('Password must be at least 6 characters.', 'error')
        return redirect(url_for('admin_users'))
    users = load_users()
    if username not in users:
        flash('User not found.', 'error')
        return redirect(url_for('admin_users'))
    users[username]['password'] = hash_password(new_pw)
    save_users(users)
    flash(f"Password for {username} has been reset.", 'success')
    return redirect(url_for('admin_users'))

@app.route('/forgot-password', methods=['GET', 'POST'])
def forgot_password():
    """Step 1 – user enters their email to request a reset link."""
    if request.method == 'POST':
        email = request.form.get('email', '').strip().lower()
        if not email:
            flash('Please enter your email address.', 'error')
            return redirect(url_for('forgot_password'))

        # Find matching user (or admin)
        target_user = None
        if ADMIN_EMAIL.lower() == email:
            target_user = ADMIN_USER
        else:
            users = load_users()
            for uname, udata in users.items():
                if udata.get('email', '').lower() == email:
                    target_user = uname
                    break

        if target_user:
            token  = uuid.uuid4().hex
            expires = (datetime.datetime.utcnow() + datetime.timedelta(hours=1)).isoformat()
            tokens  = load_reset_tokens()
            tokens[token] = {'username': target_user, 'expires': expires}
            save_reset_tokens(tokens)
            reset_url = request.host_url.rstrip('/') + url_for('reset_password', token=token)
            return render_template('forgot_password.html', reset_url=reset_url,
                                   reset_user=target_user, **ctx())

        flash('No account found with that email address.', 'error')
        return redirect(url_for('forgot_password'))
    return render_template('forgot_password.html', reset_url=None, **ctx())

@app.route('/reset-password/<token>', methods=['GET', 'POST'])
def reset_password(token):
    """Step 2 – user clicks emailed link and sets a new password."""
    tokens = load_reset_tokens()
    entry  = tokens.get(token)
    if not entry:
        flash('Reset link is invalid or has already been used.', 'error')
        return redirect(url_for('forgot_password'))

    # Check expiry
    try:
        expires = datetime.datetime.fromisoformat(entry['expires'])
    except Exception:
        expires = datetime.datetime.utcnow()
    if datetime.datetime.utcnow() > expires:
        tokens.pop(token, None)
        save_reset_tokens(tokens)
        flash('Reset link has expired. Please request a new one.', 'error')
        return redirect(url_for('forgot_password'))

    if request.method == 'POST':
        new_pw  = request.form.get('new_password', '').strip()
        confirm = request.form.get('confirm_password', '').strip()
        if len(new_pw) < 6:
            flash('Password must be at least 6 characters.', 'error')
            return redirect(url_for('reset_password', token=token))
        if new_pw != confirm:
            flash('Passwords do not match.', 'error')
            return redirect(url_for('reset_password', token=token))

        username = entry['username']
        if username == ADMIN_USER:
            flash('Owner password cannot be reset this way. Contact your host provider.', 'error')
            return redirect(url_for('login'))

        users = load_users()
        if username in users:
            users[username]['password'] = hash_password(new_pw)
            save_users(users)

        # Burn the token
        tokens.pop(token, None)
        save_reset_tokens(tokens)

        flash('Password updated! You can now log in.', 'success')
        return redirect(url_for('login'))

    return render_template('reset_password.html', token=token, **ctx())

# Keep /change-password as alias pointing to forgot flow for nav links
@app.route('/change-password')
def change_password():
    return redirect(url_for('forgot_password'))

# ── Admin – Backups ───────────────────────────────────────────────────────────
@app.route('/admin/backups')
@login_required
@admin_required
def admin_backups():
    os.makedirs(BACKUP_FOLDER, exist_ok=True)
    files = sorted(os.listdir(BACKUP_FOLDER), reverse=True)
    backups = []
    for f in files:
        if f.endswith('.csv'):
            path = os.path.join(BACKUP_FOLDER, f)
            stat = os.stat(path)
            backups.append({
                'filename': f,
                'size':     stat.st_size,
                'modified': datetime.datetime.fromtimestamp(stat.st_mtime).strftime('%Y-%m-%d %H:%M')
            })
    return render_template('admin_backups.html', backups=backups, **ctx())

@app.route('/admin/backups/download/<filename>')
@login_required
@admin_required
def download_backup(filename):
    return send_from_directory(BACKUP_FOLDER, filename, as_attachment=True)

@app.route('/admin/backups/restore/<filename>', methods=['POST'])
@login_required
@admin_required
def restore_backup(filename):
    src = os.path.join(BACKUP_FOLDER, filename)
    if os.path.exists(src):
        shutil.copy2(src, INVENTORY_FILE)
        flash(f'Inventory restored from {filename}!', 'success')
    return redirect(url_for('admin_backups'))

@app.route('/admin/backups/manual', methods=['POST'])
@login_required
@admin_required
def manual_backup():
    _backup_inventory()
    flash('Manual backup created!', 'success')
    return redirect(url_for('admin_backups'))

# ── Debug ─────────────────────────────────────────────────────────────────────
@app.route('/debug')
@login_required
@admin_required
def debug():
    info = {
        'store_name':       STORE_NAME,
        'base_dir':         BASE_DIR,
        'inventory_file':   INVENTORY_FILE,
        'inventory_exists': os.path.exists(INVENTORY_FILE),
        'upload_folder':    UPLOAD_FOLDER,
        'anthropic_key_set':bool(os.environ.get('ANTHROPIC_API_KEY')),
        'demo_mode':        DEMO_MODE,
        'python_version':   __import__('sys').version,
    }
    return jsonify(info)

# ── Context processor ─────────────────────────────────────────────────────────
@app.context_processor
def inject_globals():
    return dict(
        store_name=STORE_NAME,
        demo_mode=DEMO_MODE,
        demo_contact_email=CONTACT_EMAIL,
        stats=get_stats(),
        sale_state=load_sale(),
    )

if __name__ == '__main__':
    app.run(debug=True)
